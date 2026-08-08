"""
Tests for the ``spreadsheet_operations`` skill (data_analytics).

``create``/``edit`` run for real via openpyxl against the temp workspace
(``FORGE_WORKSPACE`` -> ``tmp_path``); output is schema-validated. The missing-
backend case is exercised by monkeypatching the module's openpyxl flag and
asserts NO file is faked. Every async test is marked (pytest-asyncio strict).
"""

from __future__ import annotations

import pytest

import skills
from skills import registry as R
from skills.data_analytics.spreadsheet_operations import handler as xlsx
from skills.data_analytics.spreadsheet_operations.handler import handle
from tests._skill_helpers import ScriptedProvider

XLSX_PATH = "reports/budget.xlsx"


@pytest.fixture
def ws(tmp_path, monkeypatch):
    (tmp_path / "reports").mkdir()
    monkeypatch.setenv("FORGE_WORKSPACE", str(tmp_path))
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr("tools._paths.FORGE_WORKSPACE", tmp_path)
    return tmp_path


def _manifest():
    return skills.get_skill("spreadsheet_operations").manifest


def _create_input():
    return {
        "operation": "create",
        "file_path": XLSX_PATH,
        "data": {"sheet": "Budget", "rows": [["Item", "Cost"], ["Infra", 4200]]},
        "formulas": ["=SUM(B2:B10)"],
    }


# 1. Input validation ---------------------------------------------------------


@pytest.mark.asyncio
async def test_missing_required_fields_raise_input_error(ws):
    with pytest.raises(R.SkillInputError):
        await handle({}, ScriptedProvider())
    with pytest.raises(R.SkillInputError):
        await handle({"operation": "create"}, ScriptedProvider())  # no file_path


@pytest.mark.asyncio
async def test_bad_operation_raises_and_no_file_created(ws):
    with pytest.raises(R.SkillInputError):
        await handle({**_create_input(), "operation": "bogus"}, ScriptedProvider())
    assert not (ws / "reports" / "budget.xlsx").exists()


# 2. Real create + real read-back ----------------------------------------------


@pytest.mark.asyncio
async def test_create_writes_real_xlsx_and_schema_valid(ws):
    out = await handle(_create_input(), ScriptedProvider())
    assert R.validate_schema(_manifest().output_schema, out) == []
    # The file was really created on disk in the workspace.
    target = ws / "reports" / "budget.xlsx"
    assert target.exists() and target.stat().st_size > 0
    # Summary reports the creation honestly.
    assert out["summary"]["operation"] == "create"
    assert out["summary"]["created"] is True
    assert out["summary"]["data_rows"] == 2
    assert out["summary"]["formulas_applied"] == 1
    assert out["charts"] == []  # never faked
    # Read the real workbook back and verify the contents are genuinely there.
    import openpyxl

    wb = openpyxl.load_workbook(str(target))
    wsheet = wb["Budget"]
    assert wsheet["A1"].value == "Item"
    assert wsheet["B2"].value == 4200
    # The formula string really landed as a formula cell in the final row.
    assert wsheet["A3"].value == "=SUM(B2:B10)"


@pytest.mark.asyncio
async def test_edit_appends_rows_to_existing_workbook(ws):
    await handle(_create_input(), ScriptedProvider())
    edit = {
        "operation": "edit",
        "file_path": XLSX_PATH,
        "data": {"rows": [["Marketing", 1500]]},
    }
    out = await handle(edit, ScriptedProvider())
    assert R.validate_schema(_manifest().output_schema, out) == []
    assert out["summary"]["operation"] == "edit"
    assert out["summary"]["created"] is False
    assert out["summary"]["data_rows_appended"] == 1
    # Verify against the real file: the appended row is present.
    import openpyxl

    wb = openpyxl.load_workbook(str(ws / "reports" / "budget.xlsx"))
    wsheet = wb.active
    values = [row[0].value for row in wsheet.iter_rows()]
    assert "Marketing" in values


@pytest.mark.asyncio
async def test_analyze_reports_real_sheet_dimensions(ws):
    await handle(_create_input(), ScriptedProvider())
    out = await handle(
        {"operation": "analyze", "file_path": XLSX_PATH}, ScriptedProvider()
    )
    assert R.validate_schema(_manifest().output_schema, out) == []
    s = out["summary"]
    assert s["operation"] == "analyze"
    assert s["sheet_count"] == 1
    assert "Budget" in s["sheets"]
    # Real dimensions from the real file (header + data + formula row = 3 rows).
    assert s["sheets"]["Budget"]["max_row"] == 3
    assert s["total_rows"] == 3


# 3. Missing-backend honesty: SkillError, no faked file -------------------------


@pytest.mark.asyncio
async def test_missing_openpyxl_raises_skillerror_and_never_fakes(ws, monkeypatch):
    monkeypatch.setattr(xlsx, "openpyxl", None)  # simulate absent backend
    with pytest.raises(R.SkillError) as excinfo:
        await handle(_create_input(), ScriptedProvider())
    assert "openpyxl" in str(excinfo.value)
    assert not (ws / "reports" / "budget.xlsx").exists()


# 4. Executor path --------------------------------------------------------------


@pytest.mark.asyncio
async def test_execute_skill_end_to_end(ws):
    out = await skills.execute_skill(
        "spreadsheet_operations", _create_input(), ScriptedProvider()
    )
    assert R.validate_schema(_manifest().output_schema, out) == []
    assert (ws / "reports" / "budget.xlsx").exists()
