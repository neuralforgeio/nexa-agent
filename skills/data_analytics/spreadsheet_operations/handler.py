# SPDX-License-Identifier: MIT
"""Skill: spreadsheet_operations.

Purpose: create, edit, or analyse an ``.xlsx`` spreadsheet in the workspace.
``create`` writes a new workbook from the supplied ``data`` rows (with optional
``formulas`` appended as a final row of real formula strings); ``edit`` appends
rows to an existing workbook; ``analyze`` reads the workbook back and reports
sheet names, dimensions, and row counts.

Permissions: ``filesystem:workspace:write`` (via ``agent.tool_api``).

Honest note: this skill uses the real ``openpyxl`` library. If ``openpyxl`` is
not installed the handler raises :class:`skills.registry.SkillError` stating
the backend is missing — it does NOT fabricate a file. ``charts`` is always
``[]`` because charts are never actually created (documented in the manifest
description). ``tool_api.write_workspace_file`` only accepts ``str``, so binary
workbook bytes are written via ``Path`` rooted at ``tool_api.workspace_path``.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from agent import tool_api
from skills._common import as_list, coerce_str, require
from skills.registry import SkillError, SkillInputError

__all__ = ["handle"]

try:
    import openpyxl
    from openpyxl.utils import get_column_letter  # noqa: F401
except ImportError:  # pragma: no cover - environment dependent
    openpyxl = None  # type: ignore[assignment]

OPENPYXL_AVAILABLE = openpyxl is not None


def _resolve(file_path: str) -> Path:
    try:
        p = tool_api.workspace_path(file_path)
    except ValueError as exc:
        raise SkillInputError(f"invalid file_path {file_path!r}: {exc}") from exc
    return Path(p)


def _require_openpyxl() -> None:
    if openpyxl is None:
        raise SkillError(
            "spreadsheet_operations requires the 'openpyxl' backend, which is "
            "not installed in this environment. Install it (pip install openpyxl) "
            "to create/edit/analyze real .xlsx workbooks; no file was faked."
        )


def _rows_from_data(data: Dict[str, Any]) -> List[List[Any]]:
    """Pull a 2D rows array out of the flexible ``data`` object."""
    if isinstance(data, dict):
        rows = data.get("rows")
        if isinstance(rows, list):
            return [list(r) if isinstance(r, (list, tuple)) else [r] for r in rows]
    return []


def _sheet_name(data: Dict[str, Any], default: str = "Sheet1") -> str:
    if isinstance(data, dict):
        return coerce_str(data.get("sheet"), default=default) or default
    return default


async def handle(input_data: dict, provider) -> dict:  # noqa: ARG001 - LLM not used
    """Run a real spreadsheet operation; no LLM is involved."""
    operation = require(input_data, "operation", str, "create|edit|analyze")
    file_path = require(input_data, "file_path", str, "workspace path of the .xlsx file")
    data = input_data.get("data") or {}
    formulas = [coerce_str(f) for f in as_list(input_data.get("formulas"))]

    if operation not in ("create", "edit", "analyze"):
        raise SkillInputError(
            f"operation must be one of create|edit|analyze, got {operation!r}"
        )

    _require_openpyxl()
    p = _resolve(file_path)

    if operation == "create":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _sheet_name(data)
        rows = _rows_from_data(data)
        for r in rows:
            ws.append(r)
        # Apply formulas as a real final row (stored as formulas, not values).
        if formulas:
            ws.append(formulas)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return {
            "file_path": file_path,
            "summary": {
                "operation": "create",
                "sheet": ws.title,
                "rows_written": len(rows) + (1 if formulas else 0),
                "data_rows": len(rows),
                "formulas_applied": len(formulas),
                "created": True,
            },
            "charts": [],
        }

    if operation == "edit":
        if not p.exists() or not p.is_file():
            raise SkillInputError(
                f"file_path {file_path!r} does not exist for edit "
                f"(resolved to {p})"
            )
        wb = openpyxl.load_workbook(str(p))
        ws = wb.active
        rows = _rows_from_data(data)
        for r in rows:
            ws.append(r)
        if formulas:
            ws.append(formulas)
        wb.save(str(p))
        return {
            "file_path": file_path,
            "summary": {
                "operation": "edit",
                "sheet": ws.title,
                "rows_appended": len(rows) + (1 if formulas else 0),
                "data_rows_appended": len(rows),
                "formulas_applied": len(formulas),
                "created": False,
            },
            "charts": [],
        }

    # operation == "analyze"
    if not p.exists() or not p.is_file():
        raise SkillInputError(
            f"file_path {file_path!r} does not exist for analyze "
            f"(resolved to {p})"
        )
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    sheets: Dict[str, Any] = {}
    total_rows = 0
    for name in wb.sheetnames:
        ws = wb[name]
        dims = ws.calculate_dimension()
        rows_count = ws.max_row or 0
        total_rows += rows_count
        sheets[name] = {
            "max_row": rows_count,
            "max_column": ws.max_column or 0,
            "dimension": dims,
        }
    wb.close()
    return {
        "file_path": file_path,
        "summary": {
            "operation": "analyze",
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
            "total_rows": total_rows,
        },
        "charts": [],
    }
